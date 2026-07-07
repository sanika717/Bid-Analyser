from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def _normalize_label(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def read_template_fields(template_path: Path) -> List[Dict[str, object]]:
    workbook = load_workbook(template_path, data_only=True)
    fields = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            values = [_normalize_label(value) for value in row if value is not None]
            if not values:
                continue
            text = " | ".join(values)
            if any(token in text.upper() for token in ["BIDDER", "DETAILS", "DATA REQUIRED"]):
                continue
            if len(values) >= 2 and any(token.lower() in {"net worth", "annual turnover", "net profit", "emd", "power of attorney", "rfq acceptance", "price", "delivery", "warranty", "payment terms", "taxes"} for token in values):
                fields.append({"sheet": sheet.title, "row": row})
    return fields


def extract_template_requirements(template_path: Path) -> List[str]:
    workbook = load_workbook(template_path, data_only=True)
    requirements: List[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            values = [_normalize_label(value) for value in row if value is not None]
            if not values:
                continue
            normalized = [value for value in values if value]
            if normalized and any(token.lower() in {"net worth", "annual turnover", "net profit", "emd", "power of attorney", "rfq acceptance", "price", "delivery", "warranty", "payment terms", "taxes"} for token in normalized):
                requirements.extend(normalized)
    cleaned = []
    for req in requirements:
        if not req:
            continue
        lowered = req.lower()
        if lowered in {"bid analysis - commercial", "bid analysis - technical", "client :", "project :"}:
            continue
        if lowered not in {item.lower() for item in cleaned}:
            cleaned.append(req)
    return cleaned
