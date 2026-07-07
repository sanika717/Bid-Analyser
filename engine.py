

    # =========================================================
# FINAL AI BID ENGINE (STABLE + IMPROVED EXTRACTION)
# =========================================================

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
from typing import Dict
import re, html, subprocess
import numpy as np

from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from PyPDF2 import PdfReader

# =========================================================
# CONFIG
# =========================================================
PROJECT_DIR = Path(__file__).parent
PDF_DIR = PROJECT_DIR / "pdfs"
TEMPLATE = PROJECT_DIR / "templates" / "template.xlsx"
OUTPUT_DIR = PROJECT_DIR / "output"

HEADER_VENDOR_START_COL = 6

MODEL_NAME = "all-MiniLM-L6-v2"
OLLAMA_MODEL = "qwen2.5:1.5b"   # ✅ use lightweight model

EMBED_MODEL = None


# =========================================================
# INIT
# =========================================================
def init_embedding():
    global EMBED_MODEL
    if EMBED_MODEL is None:
        EMBED_MODEL = SentenceTransformer(MODEL_NAME)


# =========================================================
# DATA MODEL
# =========================================================
@dataclass
class VendorProfile:
    vendor: str
    commercial: Dict[str, str] = field(default_factory=dict)


# =========================================================
# BASIC HELPERS
# =========================================================
def clean(x):
    if not x:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(x))).strip()

def norm(x):
    return clean(x).lower()

def short(x, n=200):
    return clean(x)[:n]

def clean_excel_safe(text):
    if not text:
        return ""
    text = re.sub(r"[\x00-\x1F]", "", str(text))
    return re.sub(r"\s+", " ", text).strip()


# =========================================================
# PDF READER
# =========================================================
def read_pdf(path):
    reader = PdfReader(str(path))
    text = ""
    for p in reader.pages:
        try:
            text += p.extract_text() + "\n"
        except:
            pass
    return text


# =========================================================
# CHUNKING (FIXED)
# =========================================================
def chunk_text(text, size=500, overlap=100):
    words = text.split()
    chunks = []

    for i in range(0, len(words), size - overlap):
        chunks.append(" ".join(words[i:i + size]))

    return chunks


# =========================================================
# EMBEDDINGS
# =========================================================
def build_index(text):
    chunks = chunk_text(text)
    emb = EMBED_MODEL.encode(chunks, normalize_embeddings=True)
    return chunks, emb

def semantic_search(query, chunks, emb, k=5):
    q = EMBED_MODEL.encode([query], normalize_embeddings=True)
    scores = np.dot(emb, q.T).squeeze()
    idx = np.argsort(scores)[-k:][::-1]
    return [chunks[i] for i in idx]


# =========================================================
# OLLAMA CALL
# =========================================================
def call_llm(prompt):
    try:
        res = subprocess.run(
            ["ollama", "run", OLLAMA_MODEL],
            input=prompt.encode("utf-8"),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=60
        )
        return res.stdout.decode("utf-8", errors="ignore").strip()
    except Exception:
        return ""


# =========================================================
# IMPROVED EXTRACTION (MULTI-CHUNK + VOTING)
# =========================================================
def extract_field(query, chunks, emb):

    top_chunks = semantic_search(query, chunks, emb)

    answers = []

    for c in top_chunks:

        prompt = f"""
You are extracting data from an industrial quotation.

Return ONLY the answer.
NO explanation.

If not found → return NOT_FOUND.

Field: {query}

TEXT:
{c}
"""

        res = call_llm(prompt)
        res = clean_excel_safe(res)

        if res and res.lower() != "not_found" and len(res) > 3:
            answers.append(res)

    if answers:
        # ✅ majority voting
        return max(set(answers), key=answers.count)

    return ""


# =========================================================
# SIMPLE REGEX BACKUP
# =========================================================
def regex_price(text):
    m = re.findall(r"(₹|Rs\.?|INR)\s*[0-9,]+", text)
    return m[0] if m else ""


# =========================================================
# VENDOR NAME
# =========================================================
def extract_vendor(text, filename):
    m = re.search(r"([A-Z][A-Za-z &.,]+ (Pvt\. Ltd\.|Limited|Ltd\.))", text)
    if m:
        return clean(m.group(1))
    return filename.replace(".pdf", "")


# =========================================================
# BUILD PROFILE
# =========================================================
def build_profile(pdf):

    print("📄 Processing:", pdf.name)

    text = read_pdf(pdf)
    chunks, emb = build_index(text)

    vendor = extract_vendor(text, pdf.name)
    p = VendorProfile(vendor)

    # ✅ better queries
    p.commercial["price"] = extract_field(
        "total project cost OR lump sum price OR contract price", chunks, emb
    )

    if not p.commercial["price"]:
        p.commercial["price"] = regex_price(text)

    p.commercial["delivery"] = extract_field(
        "delivery time OR delivery period OR lead time", chunks, emb
    )

    p.commercial["payment"] = extract_field(
        "payment terms OR advance payment OR milestone payment", chunks, emb
    )

    p.commercial["warranty"] = extract_field(
        "warranty period OR guarantee period", chunks, emb
    )

    return p


# =========================================================
# EXCEL HEADER DETECT
# =========================================================
def find_header_row(ws):
    for r in range(1, 20):
        txt = "".join(str(ws.cell(r, c).value).lower() for c in range(1, 10) if ws.cell(r, c).value)
        if "vendor" in txt:
            return r
    return 6


# =========================================================
# COLUMN RESET
# =========================================================
def reset_vendor_columns(ws, header, vendors):

    start = HEADER_VENDOR_START_COL

    if ws.max_column >= start:
        ws.delete_cols(start, ws.max_column - start + 1)

    col_map = {}

    for i, v in enumerate(vendors):

        col = start + i
        ws.insert_cols(col)

        cell = ws.cell(header, col)
        cell.value = v
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

        col_map[v] = col

    return col_map


# =========================================================
# EXCEL VALUE MAPPING
# =========================================================
def map_value(p, label):

    t = norm(label)

    if "price" in t:
        return p.commercial.get("price")

    if "delivery" in t:
        return p.commercial.get("delivery")

    if "payment" in t:
        return p.commercial.get("payment")

    if "warranty" in t:
        return p.commercial.get("warranty")

    return ""


# =========================================================
# FILL EXCEL
# =========================================================
def fill_workbook(wb, profiles):

    vendors = list({p.vendor: p}.values())
    vendors = [p.vendor for p in profiles]

    for sheet in wb.sheetnames:

        ws = wb[sheet]
        header = find_header_row(ws)

        col_map = reset_vendor_columns(ws, header, vendors)

        for p in profiles:
            col = col_map[p.vendor]

            for r in range(header + 1, ws.max_row):

                label = (
                    ws.cell(r, 2).value or
                    ws.cell(r, 3).value or
                    ws.cell(r, 4).value
                )

                if not label:
                    continue

                val = map_value(p, label)
                val = clean_excel_safe(val)

                ws.cell(r, col).value = short(val) if val else "-"
                ws.cell(r, col).alignment = Alignment(wrap_text=True)

    return wb


# =========================================================
# RUN
# =========================================================
def run_engine(callback=None):

    print("🚀 Engine Starting...")

    if callback:
        callback(5, "Loading embedding model...")

    init_embedding()

    if callback:
        callback(15, "Embedding model loaded")

    profiles = []

    pdfs = list(PDF_DIR.glob("*.pdf"))
    total = len(pdfs)

    if total == 0:
        raise Exception(
            f"No PDF files found in {PDF_DIR}. "
            "Place vendor PDFs inside the pdfs folder."
        )

    for i, pdf in enumerate(pdfs):

        progress = 15 + int((i / max(total, 1)) * 60)

        if callback:
            callback(progress, f"Processing {pdf.name}")

        p = build_profile(pdf)
        profiles.append(p)

        print("✅ Vendor:", p.vendor)

    if callback:
        callback(80, "Loading Excel template")

    wb = load_workbook(TEMPLATE)

    fill_workbook(wb, profiles)

    OUTPUT_DIR.mkdir(exist_ok=True)

    out = OUTPUT_DIR / (
        f"BID_OUTPUT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb.save(out)

    if callback:
        callback(100, "Completed")

    print("✅ DONE:", out)

    return {
        "output_file": str(out),
        "profiles": [
            {
                "vendor": p.vendor,
                "commercial": p.commercial
            }
            for p in profiles
        ],
        "logs": [
            f"Processed {len(profiles)} vendors",
            f"Output saved to {out}"
        ]
    }


# =========================================================
if __name__ == "__main__":
    run_engine()