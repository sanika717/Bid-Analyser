# =========================================================
# AI BID ANALYSIS ENGINE
# Structured -> Semi-structured -> LLM cascade
# Max ONE Ollama call per PDF, via HTTP API (no subprocess)
# =========================================================

from __future__ import annotations

import html
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import requests
import pdfplumber
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# =========================================================
# CONFIG
# =========================================================
PROJECT_DIR = Path(__file__).parent
PDF_DIR = PROJECT_DIR / "pdfs"
TEMPLATE = PROJECT_DIR / "templates" / "template.xlsx"
OUTPUT_DIR = PROJECT_DIR / "output"

HEADER_VENDOR_START_COL = 6

OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "qwen3:8b"
OLLAMA_TIMEOUT = 180            # seconds, qwen3:8b on M4 can take a while on first call
MAX_LLM_CHARS = 9000            # cap document text sent to the model per call

# Canonical field keys used everywhere (matches the JSON contract the LLM returns)
TEMPLATE_FIELDS = []
TEMPLATE_LABEL_PATTERNS = {}


def parse_template_fields(template_path: Path):
    """Read `template.xlsx` and build a list of parameter keys and
    label-matching regex patterns. The function inspects all sheets and
    collects labels found in columns 2-4 below the header row.
    Returns: (fields:list[str], label_patterns:dict[key->pattern])
    """
    fields = []
    label_patterns = {}
    try:
        wb = load_workbook(template_path)
    except Exception:
        return [], {}

    labels = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        hdr = find_header_row(ws)
        for r in range(hdr + 1, ws.max_row + 1):
            label = ws.cell(r, 2).value or ws.cell(r, 3).value or ws.cell(r, 4).value
            if label:
                labels.append(clean(str(label)))

    # dedupe while preserving order
    seen = set()
    for lbl in labels:
        key = re.sub(r"[^a-z0-9]+", "_", norm(lbl)).strip("_")
        if key and key not in seen:
            seen.add(key)
            fields.append(key)
            # create a flexible regex pattern from the label words
            toks = re.sub(r"[^0-9a-z ]+", " ", norm(lbl)).split()
            if toks:
                pat = r"(?:" + r"\s*|\s*".join([re.escape(t) for t in toks]) + r")"
            else:
                pat = re.escape(norm(lbl))
            label_patterns[key] = pat

    return fields, label_patterns

VENDOR_SUFFIX = re.compile(
    r"([A-Z][A-Za-z&.,'\- ]{2,80}\b"
    r"(?:Pvt\.?\s*Ltd\.?|Private Limited|Limited|Ltd\.?|LLP|Inc\.?|Corporation|Corp\.?))"
)

CURRENCY_PRICE = re.compile(r"(₹|Rs\.?|INR|\$|USD)\s?[0-9][0-9,]*(?:\.[0-9]+)?")


# =========================================================
# DATA MODEL
# =========================================================
@dataclass
class VendorProfile:
    vendor: str
    file_name: str
    doc_type: str = "unstructured"      # structured | semi_structured | unstructured | error
    confidence: str = "medium"          # high | medium | low
    commercial: Dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "vendor": self.vendor,
            "file_name": self.file_name,
            "doc_type": self.doc_type,
            "confidence": self.confidence,
            "commercial": self.commercial,
        }


# =========================================================
# BASIC HELPERS
# =========================================================
def clean(x) -> str:
    if not x:
        return ""
    return re.sub(r"\s+", " ", html.unescape(str(x))).strip()


def norm(x) -> str:
    return clean(x).lower()


def short(x, n=200) -> str:
    return clean(x)[:n]


def clean_excel_safe(text) -> str:
    if not text:
        return ""
    text = re.sub(r"[\x00-\x1F]", "", str(text))
    return re.sub(r"\s+", " ", text).strip()


# =========================================================
# PDF READING
# =========================================================
def read_pdf_text(path: Path) -> str:
    """Extract text with pdfplumber first (better layout handling), fall back to PyPDF2."""
    parts: List[str] = []
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                parts.append(page.extract_text() or "")
    except Exception:
        parts = []

    text = "\n".join(parts).strip()
    if text:
        return text

    # Fallback: PyPDF2
    try:
        reader = PdfReader(str(path))
        chunks = []
        for p in reader.pages:
            try:
                chunks.append(p.extract_text() or "")
            except Exception:
                pass
        return "\n".join(chunks).strip()
    except Exception:
        return ""


def extract_tables(path: Path) -> List[list]:
    """Return every table found in the PDF as a list of row-lists."""
    tables: List[list] = []
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                for t in page.extract_tables() or []:
                    if t:
                        tables.append(t)
    except Exception:
        pass
    return tables


# =========================================================
# 1) STRUCTURED EXTRACTION (TABLES)
# =========================================================
def _value_from_row(cells: List[str], pattern: str) -> str:
    for i, c in enumerate(cells):
        if re.search(pattern, c, re.IGNORECASE):
            # same-cell "Label: Value" or "Label - Value"
            split = re.split(r":|-|–", c, maxsplit=1)
            if len(split) > 1 and clean(split[1]):
                return clean(split[1])
            # otherwise take the next non-empty cell in the row
            if i + 1 < len(cells):
                return clean(cells[i + 1])
    return ""


def extract_from_tables(tables: List[list]) -> Dict[str, str]:
    hits: Dict[str, str] = {}
    for table in tables:
        for row in table:
            cells = [clean(c) for c in row if c is not None and clean(c)]
            if not cells:
                continue
            row_text = " | ".join(cells)
            for key, pattern in TEMPLATE_LABEL_PATTERNS.items():
                if hits.get(key):
                    continue
                if re.search(pattern, row_text, re.IGNORECASE):
                    value = _value_from_row(cells, pattern)
                    if value:
                        hits[key] = short(value, 150)
    return hits


# =========================================================
# 2) SEMI-STRUCTURED EXTRACTION (LABEL REGEX ON TEXT)
# =========================================================
def extract_from_regex(text: str) -> Dict[str, str]:
    hits: Dict[str, str] = {}

    for line in text.splitlines():
        for key, pattern in TEMPLATE_LABEL_PATTERNS.items():
            if hits.get(key):
                continue
            m = re.search(pattern + r"\s*[:\-–]\s*(.+)", line, re.IGNORECASE)
            if m:
                val = clean(m.group(1))
                if val:
                    hits[key] = short(val, 150)

    # If a field resembling price exists in template, prefer it for fallback
    for k in ("price", "cost", "amount"):
        if k in TEMPLATE_FIELDS and not hits.get(k):
            m = CURRENCY_PRICE.search(text)
            if m:
                hits[k] = clean(m.group(0))
                break

    return hits


# =========================================================
# VENDOR NAME
# =========================================================
def extract_vendor_name(text: str, tables: List[list], filename: str) -> str:
    m = VENDOR_SUFFIX.search(text)
    if m:
        return clean(m.group(1))

    for table in tables[:2]:
        for row in table[:3]:
            for cell in row:
                if cell:
                    m = VENDOR_SUFFIX.search(str(cell))
                    if m:
                        return clean(m.group(1))

    return Path(filename).stem


# =========================================================
# 3) LLM EXTRACTION (OLLAMA HTTP API) — ONE CALL PER PDF, MAX
# =========================================================
def call_ollama(prompt: str, max_tokens: int = 400) -> str:
    # kept for backward compatibility; may be monkeypatched by engine package
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": 0.1, "num_predict": max_tokens},
    }
    try:
        resp = requests.post(OLLAMA_URL, json=payload, timeout=OLLAMA_TIMEOUT)
        resp.raise_for_status()
        return (resp.json().get("response") or "").strip()
    except Exception:
        return ""


def extract_json_block(text: str) -> dict:
    """Pull the first balanced {...} block out of the model's output.
    Handles qwen3 reasoning models that prepend <think>...</think> chatter
    before the actual JSON answer."""
    start = text.find("{")
    if start == -1:
        return {}

    depth = 0
    for i in range(start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                candidate = text[start:i + 1]
                try:
                    return json.loads(candidate)
                except Exception:
                    return {}
    return {}


def llm_extract(text: str, need_vendor: bool, missing_fields: List[str]) -> dict:
    """Single combined extraction call. Only invoked when something is
    still missing after the structured/regex passes."""
    if not (need_vendor or missing_fields):
        return {}

    snippet = text[:MAX_LLM_CHARS]
    wanted = (["vendor_name"] if need_vendor else []) + missing_fields

    # Build JSON skeleton dynamically from TEMPLATE_FIELDS
    json_fields = {f: "" for f in TEMPLATE_FIELDS}
    # Ensure vendor_name present if requested
    if need_vendor:
        json_fields["vendor_name"] = ""

    shape = json.dumps(json_fields, indent=2)

    prompt = f"""/no_think
You are extracting bid data from a vendor quotation document.
Respond with ONLY a single valid JSON object. No explanation, no markdown fences.

Return exactly this shape (use "" for any value you cannot find):
{shape}

Only these fields are actually needed: {", ".join(wanted)}

DOCUMENT TEXT:
{snippet}
"""
    raw = call_ollama(prompt)
    return extract_json_block(raw)


# =========================================================
# BUILD ONE VENDOR PROFILE
# =========================================================
def build_profile(pdf_path: Path) -> VendorProfile:
    text = read_pdf_text(pdf_path)
    tables = extract_tables(pdf_path)

    # initialize commercial dict from parsed template fields
    commercial = {k: "" for k in TEMPLATE_FIELDS} if TEMPLATE_FIELDS else {}
    doc_type = "unstructured"

    # --- Pass 1: structured (tables) ---
    table_hits = extract_from_tables(tables)
    if table_hits:
        doc_type = "structured"
        commercial.update(table_hits)

    # --- Pass 2: semi-structured (label regex on raw text) ---
    regex_hits = extract_from_regex(text)
    for k, v in regex_hits.items():
        if not commercial.get(k) and v:
            commercial[k] = v
            if doc_type == "unstructured":
                doc_type = "semi_structured"

    vendor = extract_vendor_name(text, tables, pdf_path.name)
    need_vendor = (vendor == pdf_path.stem)  # fell back to filename -> not confident

    missing = [k for k in TEMPLATE_FIELDS if not commercial.get(k)]

    # --- Pass 3: LLM, only if something is still missing (max 1 call) ---
    if missing or need_vendor:
        result = llm_extract(text, need_vendor, missing)
        if result:
            if need_vendor and result.get("vendor_name"):
                vendor = clean(result["vendor_name"])
            for key in TEMPLATE_FIELDS:
                if not commercial.get(key) and result.get(key):
                    commercial[key] = clean_excel_safe(short(result[key], 150))
        confidence = "high" if not (missing or need_vendor) else (
            "medium" if any(commercial.values()) else "low"
        )
    else:
        confidence = "high"

    return VendorProfile(
        vendor=vendor or pdf_path.stem,
        file_name=pdf_path.name,
        doc_type=doc_type,
        confidence=confidence,
        commercial=commercial,
    )


# =========================================================
# EXCEL: HEADER DETECTION
# =========================================================
def find_header_row(ws) -> int:
    for r in range(1, 20):
        row_text = "".join(
            str(ws.cell(r, c).value).lower()
            for c in range(1, 10)
            if ws.cell(r, c).value
        )
        if "vendor" in row_text:
            return r
    return 6


# =========================================================
# EXCEL: VENDOR COLUMNS
# =========================================================
def reset_vendor_columns(ws, header: int, vendors: List[str]) -> Dict[str, int]:
    start = HEADER_VENDOR_START_COL

    if ws.max_column >= start:
        ws.delete_cols(start, ws.max_column - start + 1)

    col_map: Dict[str, int] = {}
    for i, v in enumerate(vendors):
        col = start + i
        ws.insert_cols(col)

        cell = ws.cell(header, col)
        cell.value = v
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

        col_map[v] = col

    return col_map


# =========================================================
# EXCEL: VALUE MAPPING (row label -> profile field)
# =========================================================
def map_value(p: VendorProfile, label: str) -> str:
    t = norm(label)
    # Try to match the label against template label patterns
    for key, pat in TEMPLATE_LABEL_PATTERNS.items():
        try:
            if re.search(pat, t, re.IGNORECASE):
                return p.commercial.get(key, "")
        except re.error:
            continue
    # fallback: attempt naive keyword matching
    for key in TEMPLATE_FIELDS:
        if key in t:
            return p.commercial.get(key, "")
    return ""


# =========================================================
# EXCEL: FILL WORKBOOK
# =========================================================
def fill_workbook(wb, profiles: List[VendorProfile]):
    # de-dupe by vendor name while preserving order
    seen = set()
    vendors = []
    for p in profiles:
        if p.vendor not in seen:
            seen.add(p.vendor)
            vendors.append(p.vendor)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = find_header_row(ws)
        col_map = reset_vendor_columns(ws, header, vendors)

        for p in profiles:
            col = col_map.get(p.vendor)
            if col is None:
                continue

            # NOTE: range stop is +1 so the last row is included (off-by-one
            # in the original script meant the bottom row never got filled).
            for r in range(header + 1, ws.max_row + 1):
                label = (
                    ws.cell(r, 2).value
                    or ws.cell(r, 3).value
                    or ws.cell(r, 4).value
                )
                if not label:
                    continue

                val = clean_excel_safe(map_value(p, label))
                ws.cell(r, col).value = short(val) if val else "-"
                ws.cell(r, col).alignment = Alignment(wrap_text=True)

    return wb


# =========================================================
# MAIN ENTRY POINT
# =========================================================
def run_engine(callback=None) -> dict:
    """
    callback(percent: int, message: str) is called repeatedly to report progress.
    Returns: {"output_file": str, "profiles": [dict, ...], "logs": [str, ...]}
    """
    logs: List[str] = []

    def log(pct: int, msg: str):
        logs.append(f"{pct}% - {msg}")
        if callback:
            callback(pct, msg)

    log(2, "Starting engine...")

    # Parse template-driven fields/patterns so extraction is dynamic
    global TEMPLATE_FIELDS, TEMPLATE_LABEL_PATTERNS
    TEMPLATE_FIELDS, TEMPLATE_LABEL_PATTERNS = parse_template_fields(TEMPLATE)

    PDF_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(PDF_DIR.glob("*.pdf"))

    if not pdf_files:
        log(100, "No PDFs found in pdfs/ folder")
        wb = load_workbook(TEMPLATE)
        out_path = OUTPUT_DIR / f"BID_OUTPUT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(out_path)
        return {"output_file": str(out_path), "profiles": [], "logs": logs}

    total = len(pdf_files)
    log(5, f"Found {total} PDF(s) to process")

    profiles: List[VendorProfile] = []
    span = 70.0  # progress budget for the PDF loop: 5% -> 75%

    for i, pdf in enumerate(pdf_files):
        pct = int(5 + (i / total) * span)
        log(pct, f"Reading {pdf.name} ({i + 1}/{total})")

        try:
            profile = build_profile(pdf)
            log(int(5 + ((i + 1) / total) * span),
                f"Extracted vendor '{profile.vendor}' from {pdf.name} [{profile.doc_type}]")
        except Exception as e:
            profile = VendorProfile(
                vendor=pdf.stem,
                file_name=pdf.name,
                doc_type="error",
                confidence="low",
            )
            log(pct, f"Failed to process {pdf.name}: {e}")

        profiles.append(profile)

    log(80, "Generating Excel comparison sheet...")
    wb = load_workbook(TEMPLATE)
    fill_workbook(wb, profiles)

    log(92, "Saving Excel file...")
    out_path = OUTPUT_DIR / f"BID_OUTPUT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)

    log(100, "Completed")

    return {
        "output_file": str(out_path),
        "profiles": [p.to_dict() for p in profiles],
        "logs": logs,
    }


if __name__ == "__main__":
    run_engine()