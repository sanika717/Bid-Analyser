from __future__ import annotations

from pathlib import Path
from typing import List, Dict

from openpyxl import Workbook


def write_workbook(output_path: Path, profiles: List[Dict], ranked: List[Dict]) -> Path:
    workbook = Workbook()
    comparison_sheet = workbook.active
    comparison_sheet.title = "Bid Comparison"
    headers = ["Vendor", "File", "Doc Type", "Confidence"]
    fields = sorted({key for profile in profiles for key in profile.get("commercial", {}).keys()})
    headers.extend(fields)
    comparison_sheet.append(headers)
    for profile in profiles:
        row = [profile.get("vendor"), profile.get("file_name"), profile.get("doc_type"), profile.get("confidence")]
        commercial = profile.get("commercial", {})
        for field in fields:
            row.append(commercial.get(field, ""))
        comparison_sheet.append(row)

    ranking_sheet = workbook.create_sheet("Vendor Ranking")
    ranking_sheet.append(["Vendor", "File", "Score", "Confidence"])
    for item in ranked:
        ranking_sheet.append([item.get("vendor"), item.get("file_name"), item.get("score"), item.get("confidence")])

    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_sheet.append(["Vendor", "Evidence"])
    for profile in profiles:
        evidence_text = []
        for key, value in profile.get("commercial", {}).items():
            if value:
                evidence_text.append(f"{key}: {value}")
        evidence_sheet.append([profile.get("vendor"), " | ".join(evidence_text)])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
